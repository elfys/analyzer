from datetime import (
    date,
    datetime,
)
from decimal import Decimal
from typing import (
    Iterable,
    Sequence,
    TypedDict,
)

import click
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from sqlalchemy.orm import (
    Query,
    contains_eager,
    joinedload,
    undefer,
)

from orm import (
    Chip,
    ChipState,
    IVMeasurement,
    IvConditions,
    Matrix,
    MatrixChip,
    Wafer,
)
from utils import (
    EntityOption,
    get_indexed_filename,
    get_thresholds,
    validate_chip_types,
    wafer_loader,
)
from .common import (
    apply_conditional_formatting,
    date_formats,
    date_formats_help,
    get_info,
    get_slice_by_voltages,
    plot_data,
)
from ..context import (
    AnalyzerContext,
    pass_analyzer_context,
)


class SheetsIVData[T](TypedDict):
    anode: T
    cathode: T
    guard_ring: T
    anode_raw: T
    temperatures: T


@click.command(name="iv", help="Make summary (png and xlsx) for IV measurements' data.")
@pass_analyzer_context
@click.option(
    "-t", "--chips-type", help="Type of the chips to analyze.",
    callback=validate_chip_types,
)
@click.option(
    "-w", "--wafer", prompt=True, help="Wafer name.", required=True,
    callback=wafer_loader
)
@click.option(
    "-s",
    "--chip-state",
    "chip_states",
    help="State of the chips to summarize.",
    default=["ALL"],
    show_default=True,
    multiple=True,
    cls=EntityOption,
    entity_type=ChipState,
)
@click.option(
    "-q",
    "--quantile",
    default=(0.01, 0.99),
    show_default=True,
    type=click.Tuple([float, float]),
    help="Min and max plotted values cutoff.",
)
@click.option(
    "--before",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements before (exclusive) provided date and time. {date_formats_help}",
)
@click.option(
    "--after",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements after (inclusive) provided date and time. {date_formats_help}",
)
def summary_iv(
    ctx: AnalyzerContext,
    chips_type: str | None,
    wafer: Wafer,
    chip_states: Sequence[ChipState],
    quantile: tuple[float, float],
    before: datetime | date | None,
    after: datetime | date | None,
):
    query: Query = (
        ctx.session.query(IvConditions)
        .join(IvConditions.measurements)
        .filter(
            IvConditions.chip.has(Chip.wafer == wafer),
            IvConditions.chip_state_id.in_((c.id for c in chip_states)),
        )
        .options(
            contains_eager(IvConditions.measurements),
            joinedload(IvConditions.chip),
            undefer(IvConditions.datetime),
        )
    )
    
    if chips_type:
        query = query.filter(IvConditions.chip.has(Chip.type == chips_type))
    else:
        ctx.logger.info(
            "Chips type (-t or --chips-type) is not specified. Analyzing all chip types."
        )
    
    if before is not None or after is not None:
        after = after if after is not None else date.min
        before = before if before is not None else date.max
        query = query.filter(IvConditions.datetime.between(after, before))
    
    conditions: list[IvConditions] = query.all()
    
    if not conditions:
        ctx.logger.warning("No measurements found.")
        return
    
    chips = {c.chip for c in conditions}
    measurements = get_iv_measurements(conditions)
    sheets_data = get_sheets_iv_data(measurements, chips)
    
    summary_voltages = list(sheets_data["anode"].columns.intersection(
        [Decimal(v) for v in {"-1", "0.01", "5", "6", "10", "20"}]
    ))
    thresholds = get_thresholds(ctx.session, "IV")
    
    chips_types = {chips_type} if chips_type else {chip.type for chip in chips}
    
    matrix_chips = [chip for chip in chips if isinstance(chip, MatrixChip)]
    if matrix_chips:
        # fetch matrices to avoid multiple queries later
        matrices = ctx.session.query(Matrix) \
            .filter(Matrix.id.in_([c.matrix_id for c in matrix_chips])) \
            .all()
    
    title = f"{wafer.name} {','.join(chips_types)}"
    file_name = get_indexed_filename(f"Summary-IV-{title.replace(' ', '-')}", ("png", "xlsx"))
    
    if len(chips_types) > 1:
        ctx.logger.warning(
            f"Multiple chip types are found ({chips_types}). "
            "Plotting is not supported and will be skipped.")
    else:
        chips_type = next(iter(chips_types))
        fig, axes = plot_data(
            measurements,
            summary_voltages,
            quantile,
            thresholds.get(chips_type, {}),
        )
        fig.suptitle(title, fontsize=14)
        
        png_file_name = f"{file_name}.png"
        fig.savefig(png_file_name, dpi=300)
        ctx.logger.info(f"Summary data is plotted to {png_file_name}")
    
    exel_file_name = f"{file_name}.xlsx"
    info = get_info(wafer=wafer, chip_states=chip_states, measurements=measurements)
    save_iv_summary_to_excel(sheets_data, info, exel_file_name, summary_voltages, thresholds)
    
    ctx.logger.info(f"Summary data is saved to {exel_file_name}")


def get_iv_measurements(conditions: list[IvConditions]) -> list[IVMeasurement]:
    # sort conditions by amplitude of voltage, smaller go last
    # thus more precise measurements with voltage input from -0.01 to 0.01 will overwrite less
    # precise from -1 to 20
    
    def get_sort_keys(condition: IvConditions):
        voltages = [m.voltage_input for m in condition.measurements]
        voltage_amplitude = max(voltages) - min(voltages)
        return condition.datetime, -voltage_amplitude
    
    non_empty_conditions = [c for c in conditions if c.measurements]
    sorted_conditions = list(sorted(non_empty_conditions, key=get_sort_keys, reverse=False))
    
    # get all measurements, deduplicated by voltage and chip name
    # latest measurements from `sorted_conditions` will be selected
    measurements = (
        (m.voltage_input, c.chip_id, m) for c in sorted_conditions for m in c.measurements)
    measurements_dict = {(v, c): m for v, c, m in measurements}
    return list(measurements_dict.values())


def save_iv_summary_to_excel(
    sheets_data: SheetsIVData[pd.DataFrame],
    info: pd.Series,
    file_name: str,
    voltages: Iterable[Decimal],
    thresholds: dict[str, dict[Decimal, float]],
):
    summary_df = get_slice_by_voltages(sheets_data["anode"], voltages)
    summary_df.insert(
        0, "Temperature", sheets_data["temperatures"]["Temperature"].apply(lambda x: f"{x:.2f}"))
    rules = {
        "lessThan": PatternFill(bgColor="ee9090", fill_type="solid"),  # red, failed
        "greaterThanOrEqual": PatternFill(bgColor="90ee90", fill_type="solid"),  # green, ok
    }
    sheet_names: SheetsIVData[str | None] = {
        'anode': "I1 anode",
        'cathode': "I3 anode",
        'guard_ring': "Guard ring current",
        'anode_raw': "I1 anode raw",
        'temperatures': None,
    }
    
    with pd.ExcelWriter(file_name) as writer:
        summary_df.to_excel(writer, sheet_name="Summary")
        apply_conditional_formatting(
            writer.book["Summary"],
            rules,
            thresholds,
        )
        for df_name, df in sheets_data.items():
            if (sheet_name := sheet_names[df_name]) is None:
                continue
            df = df.dropna(axis=1, how="all")
            if df.empty:
                continue
            df = df.rename(columns=float)
            df.to_excel(writer, sheet_name=sheet_name)
        
        info.to_excel(writer, sheet_name="Info")


@pass_analyzer_context
def get_sheets_iv_data(
    ctx: AnalyzerContext,
    measurements: Iterable[IVMeasurement],
    chips: Iterable[Chip],
) -> SheetsIVData[pd.DataFrame]:
    chip_names = sorted({c.name for c in chips})
    temps_df = pd.DataFrame(
        0, dtype="float32", index=chip_names, columns=['Temperature', 'num_of_measurements'])
    empty_df = pd.DataFrame(dtype="float32", index=chip_names)
    anode_df = empty_df.copy()
    raw_anode_df = empty_df.copy()
    cathode_df = empty_df.copy()
    guard_ring_df = empty_df.copy()
    has_uncorrected_current = False
    with click.progressbar(measurements, label="Processing measurements...") as progress:
        for measurement in progress:
            cell_location = (measurement.conditions.chip.name, measurement.voltage_input)
            if measurement.anode_current_corrected is None:
                has_uncorrected_current = True
            anode_df.loc[cell_location] = measurement.get_anode_current_value()
            raw_anode_df.loc[cell_location] = measurement.anode_current
            cathode_df.loc[cell_location] = measurement.cathode_current
            guard_ring_df.loc[cell_location] = measurement.guard_current
            
            temps_df.loc[cell_location[0], ['Temperature', 'num_of_measurements']] += \
                np.array([measurement.conditions.temperature, 1.0], dtype=np.float32)
    if has_uncorrected_current:
        ctx.logger.warning(
            "Some current measurements are not corrected by temperature."
        )
    temps_df['Temperature'] /= temps_df['num_of_measurements']
    temps_df.drop(columns='num_of_measurements', inplace=True)
    return {
        "anode": anode_df,
        "anode_raw": raw_anode_df,
        "cathode": cathode_df,
        "guard_ring": guard_ring_df,
        "temperatures": temps_df,
    }
