from decimal import Decimal
from time import strftime
from typing import (
    Iterable,
    Sequence,
)

import click
import numpy as np
import pandas as pd
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import (
    bindparam,
    select,
)

from orm import (
    Chip,
    ChipRepository,
    ChipState,
    IVMeasurement,
    IvConditions,
    Wafer,
)
from utils import (
    EntityOption,
    get_thresholds,
    wafer_loader,
)
from .context import (
    AnalyzerContext,
    pass_analyzer_context,
)


@click.command(name="wafers", help="Compare wafers")
@pass_analyzer_context
@click.option(
    "-w",
    "--wafers",
    help="Wafers to compare",
    prompt=True,
    required=True,
    multiple=True,
    callback=wafer_loader,
)
@click.option(
    "-s",
    "--chip-state",
    "chip_states",
    help="States of chips to analyze",
    default=["ALL"],
    show_default=True,
    multiple=True,
    cls=EntityOption,
    entity_type=ChipState,
)
@click.option(
    "-o",
    "--output",
    "file_name",
    default=lambda: f"wafers-comparison-{strftime('%y%m%d-%H%M%S')}.xlsx",
    help="Output file name.",
    show_default="wafers-comparison-{datetime}.xlsx",
)
def compare_wafers(
    ctx: AnalyzerContext, wafers: set[Wafer], chip_states: Sequence[ChipState], file_name: str
):
    sheets_data = get_sheets_data(wafers)
    
    if not sheets_data["frame_keys"]:
        ctx.logger.warning("No data to compare")
        return
    
    save_compare_wafers_report(file_name, sheets_data, chip_states)
    ctx.logger.info(f"Wafers comparison is saved to {file_name}")


def save_compare_wafers_report(file_name, sheets_data, chip_states):
    chip_states_dict = {state.id: state.name for state in chip_states}
    frame_keys = sheets_data.pop("frame_keys")
    with pd.ExcelWriter(file_name) as writer:
        for key, data in sheets_data.items():
            df = pd.concat(data["frames"], keys=frame_keys, copy=False)
            df.dropna(how="all", axis=1, inplace=True)
            df.columns = pd.MultiIndex.from_tuples(
                df.columns.values, names=["Voltage", "Chip type"]
            )
            df.sort_index(axis=1, inplace=True, level=[0, 1])
            
            df.index = pd.MultiIndex.from_tuples(
                df.index.map(lambda idx: (idx[0], chip_states_dict[idx[1]])),
                names=["Wafer", "Chip state"],
            )
            if key != "yield":
                df.columns = add_perimeter_area_level(df.columns)
            
            df.to_excel(writer, sheet_name=data["title"])
            
            ws: Worksheet = writer.sheets[data["title"]]
            start_col = get_column_letter(df.index.nlevels + 1)
            start_row = df.columns.nlevels + 1
            values_slice = f"{start_col}{start_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            
            number_format = FORMAT_PERCENTAGE_00 if key == "yield" else "0.00E+00"
            for row in ws[values_slice]:
                for cell in row:
                    cell.number_format = number_format


@pass_analyzer_context
def get_sheets_data(ctx: AnalyzerContext, wafers: Iterable[Wafer]) -> dict[str, dict]:
    thresholds = get_thresholds(ctx.session, "IV")
    
    threshold_voltages = set({v for x in thresholds.values() for v in x.keys()})
    conditions_query = (
        select(IvConditions, Chip.type.label("chip_type"), IVMeasurement)
        .join(IvConditions.measurements)
        .join(IvConditions.chip)
        .filter(
            Chip.wafer_id == bindparam("wafer_id"),
            Chip.type != "TS",
            IVMeasurement.voltage_input.in_(threshold_voltages),
        )
        .order_by(IvConditions.datetime)
    )
    
    sheets_data = {
        "yield": {"frames": [], "title": "Yield"},
        "std": {"frames": [], "title": "Standard Deviation"},
        "leakage": {"frames": [], "title": "Leakage"},
        "density": {"frames": [], "title": "Density"},
        "frame_keys": [],
    }
    
    for wafer in wafers:
        values_frame = pd.read_sql_query(
            conditions_query.params(wafer_id=wafer.id), ctx.session.connection()
        )
        
        if values_frame.empty:
            ctx.logger.warning(f"Not found measurements for {wafer.name}")
            continue
        sheets_data["frame_keys"].append(wafer.name)
        
        values_frame["value"] = values_frame["anode_current_corrected"].fillna(
            values_frame["anode_current"]
        )
        
        values_frame.drop_duplicates(
            subset=["voltage_input", "chip_id", "chip_state_id"],
            keep="last",
            inplace=True,
        )
        
        values_frame = values_frame.pivot_table(
            values="value",
            columns="voltage_input",
            index=["chip_type", "chip_state_id", "chip_id"],
        )
        values_frame.rename(
            columns=lambda x: Decimal(x).quantize(next(iter(threshold_voltages))),
            inplace=True,
        )
        
        def remove_cols_wo_thresholds(frame: pd.DataFrame):
            chip_type = frame.name
            cols_to_drop = set(frame.columns.values) - set(thresholds[chip_type].keys())
            frame.loc[:, list(cols_to_drop)] = np.nan
            return frame
        
        values_frame = values_frame.groupby(["chip_type"]).apply(remove_cols_wo_thresholds)
        
        sheets_data["yield"]["frames"].append(get_yield_frame(thresholds, values_frame))
        sheets_data["std"]["frames"].append(get_std_frame(values_frame))
        sheets_data["leakage"]["frames"].append(get_leakage_frame(values_frame))
        sheets_data["density"]["frames"].append(get_density_frame(values_frame))
    
    return sheets_data


def get_density_frame(values_frame):
    def divide_by_area(frame):
        chip_type = frame.name
        return frame / ChipRepository.get_area(chip_type)
    
    density_frame = (
        values_frame.groupby("chip_type")
        .apply(divide_by_area)
        .groupby(["chip_type", "chip_state_id"])
        .median()
    )
    density_frame = density_frame.unstack(level="chip_type")
    return density_frame


def get_yield_frame(thresholds, values_frame) -> pd.DataFrame:
    row_thresholds = {
        chip_type: np.array(
            [chip_type_thresholds.get(column, np.nan) for column in values_frame.columns]
        )
        for chip_type, chip_type_thresholds in thresholds.items()
    }
    
    def pass_threshold(frame):
        chip_type = frame.name
        th: np.ndarray = row_thresholds[chip_type]
        return pd.DataFrame(
            np.where(np.isnan(frame), np.nan, frame >= th),
            columns=frame.columns,
            index=frame.index,
            copy=False,
        )
    
    pass_frame = values_frame.groupby("chip_type").apply(pass_threshold)
    total_series = pass_frame.min(axis=1).groupby("chip_state_id").mean()
    total_series.name = ("Total", "")
    yield_frame = pass_frame.groupby(["chip_type", "chip_state_id"]).mean()
    yield_frame = yield_frame.unstack(level="chip_type")
    return pd.concat(
        [
            yield_frame,
            total_series,
        ],
        axis=1,
        copy=False,
    )


def get_leakage_frame(values_frame) -> pd.DataFrame:
    leakage_frame = (
        values_frame.groupby(["chip_type", "chip_state_id"]).median().unstack(level="chip_type")
    )
    return leakage_frame


@pass_analyzer_context
def add_perimeter_area_level(ctx: AnalyzerContext, index: pd.MultiIndex) -> pd.MultiIndex:
    idx_tuples = []
    for col in index.values:
        voltage, chip_type = col
        try:
            perimeter = ChipRepository.get_perimeter(chip_type)
            area = ChipRepository.get_area(chip_type)
        except AttributeError:
            ctx.logger.warning(f"Chip {chip_type} has no perimeter or area")
            continue
        perimeter_area = perimeter / area
        idx_tuples.append((voltage, chip_type, Decimal(perimeter_area).quantize(Decimal("0.001"))))
    return pd.MultiIndex.from_tuples(idx_tuples, names=[*index.names, "Perimeter / area"])


def get_std_frame(values_frame) -> pd.DataFrame:
    std_frame = (
        values_frame.groupby(["chip_type", "chip_state_id"]).std().unstack(level="chip_type")
    )
    return std_frame


@click.group(
    name="compare",
    help="Set of commands to compare entities",
    commands=[compare_wafers],
)
def compare_group():
    ...
