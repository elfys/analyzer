from datetime import datetime, date
from decimal import Decimal
from itertools import chain
from pathlib import Path
from time import strftime, localtime
from typing import Union, Any, Iterable, Optional, Collection

import click
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from matplotlib.ticker import MaxNLocator
from numpy import ndarray
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Fill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import desc, text
from sqlalchemy.orm import Session, joinedload, contains_eager, undefer

from orm import (
    CVMeasurement,
    Chip,
    ChipState,
    EqeConditions,
    EqeSession,
    IVMeasurement,
    IvConditions,
    Wafer,
)
from utils import (
    eqe_session_date,
    EntityOption,
    get_thresholds,
)

date_formats = ["%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"]
date_formats_help = f"Supported formats are: {', '.join((strftime(f) for f in date_formats))}."


@click.command(name="iv", help="Make summary (png and xlsx) for IV measurements' data.")
@click.pass_context
@click.option("-t", "--chips-type", help="Type of the chips to analyze.")
@click.option("-w", "--wafer", "wafer_name", prompt=True, help="Wafer name.")
@click.option(
    "-o",
    "--output",
    "file_name",
    show_default="Summary-IV-{wafer}",
    help="Output file name without extension.",
)
@click.option(
    "-s",
    "--chip-state",
    "chip_states",
    help="State of the chips to summarize.",
    default=["ALL"],
    show_default=True,
    multiple=True,
    cls=EntityOption,
    entity_type=ChipState,
)
@click.option(
    "-q",
    "--quantile",
    default=(0.01, 0.99),
    show_default=True,
    type=click.Tuple([float, float]),
    help="Min and max plotted values cutoff.",
)
@click.option(
    "--before",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements before (exclusive) provided date and time. {date_formats_help}",
)
@click.option(
    "--after",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements after (inclusive) provided date and time. {date_formats_help}",
)
def summary_iv(
    ctx: click.Context,
    chips_type: Optional[str],
    wafer_name: str,
    file_name: Optional[str],
    chip_states: tuple[ChipState],
    quantile: tuple[float, float],
    before: Optional[datetime],
    after: Optional[datetime],
):
    session: Session = ctx.obj["session"]
    if ctx.obj["default_wafer"].name != wafer_name:
        wafer = session.query(Wafer).filter(Wafer.name == wafer_name).first()
    else:
        wafer = ctx.obj["default_wafer"]

    query = (
        session.query(IvConditions)
        .join(IvConditions.measurements)
        .filter(
            IvConditions.chip.has(Chip.wafer == wafer),
            IvConditions.chip_state_id.in_((c.id for c in chip_states)),
        )
        .options(
            contains_eager(IvConditions.measurements),
            joinedload(IvConditions.chip),
            undefer(IvConditions.datetime),
        )
    )

    if chips_type is not None:
        query = query.filter(IvConditions.chip.has(Chip.type == chips_type))
    else:
        ctx.obj["logger"].info(
            "Chips type (-t or --chips-type) is not specified. Analyzing all chip types."
        )

    if before is not None or after is not None:
        after = after if after is not None else date.min
        before = before if before is not None else date.max
        query = query.filter(IvConditions.datetime.between(after, before))

    conditions = query.all()

    if not conditions:
        ctx.obj["logger"].warning("No measurements found.")
        return

    sheets_data = get_sheets_iv_data(conditions)

    voltages = sheets_data["anode"].columns.intersection(
        map(Decimal, ["-1", "0.01", "5", "6", "10", "20"])
    )
    thresholds = get_thresholds(session, "IV")

    chips_types = (
        {chips_type}
        if chips_type is not None
        else {condition.chip.type for condition in conditions}
    )

    file_name = file_name or f"Summary-IV-{wafer_name}"
    file_name = get_indexed_filename(file_name, ("png", "xlsx"))

    if len(chips_types) > 1:
        ctx.obj["logger"].warning(
            f"Multiple chip types are found ({chips_types}). Plotting is not supported and will be skipped."
        )
    else:
        chips_type = next(iter(chips_types))
        fig, axes = plot_data(
            tuple(chain(*(condition.measurements for condition in conditions))),
            voltages,
            quantile,
            thresholds.get(chips_type, {}),
        )
        fig.suptitle(wafer_name, fontsize=14)
        for ax_row in axes:
            ax_row[0].set_xlabel("Anode current [pA]")

        png_file_name = f"{file_name}.png"
        fig.savefig(png_file_name, dpi=300)
        ctx.obj["logger"].info(f"Summary data is plotted to {png_file_name}")

    exel_file_name = f"{file_name}.xlsx"
    info = get_info(wafer=wafer, chip_states=chip_states, measurements=conditions)
    save_iv_summary_to_excel(sheets_data, info, exel_file_name, voltages, thresholds)

    ctx.obj["logger"].info(f"Summary data is saved to {exel_file_name}")


@click.command(name="cv", help="Make summary (png and xlsx) for CV measurements' data.")
@click.pass_context
@click.option("-t", "--chips-type", help="Type of the chips to analyze.")
@click.option("-w", "--wafer", "wafer_name", prompt="Wafer name", help="Wafer name.")
@click.option(
    "-o",
    "--output",
    "file_name",
    show_default="Summary-CV-{wafer}",
    help="Output file name without extension.",
)
@click.option(
    "-s",
    "--chip-state",
    "chip_states",
    help="State of the chips to summarize.",
    default=["ALL"],
    show_default=True,
    multiple=True,
    cls=EntityOption,
    entity_type=ChipState,
)
@click.option(
    "-q",
    "--quantile",
    default=(0.01, 0.99),
    show_default=True,
    type=click.Tuple([float, float]),
    help="Min and max plotted values cutoff.",
)
@click.option(
    "--before",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements before (exclusive) provided date and time. {date_formats_help}",
)
@click.option(
    "--after",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements after (inclusive) provided date and time. {date_formats_help}",
)
def summary_cv(
    ctx: click.Context,
    chips_type: Union[str, None],
    wafer_name: str,
    file_name: str,
    chip_states: list[ChipState],
    quantile: tuple[float, float],
    before: Union[datetime, None],
    after: Union[datetime, None],
):
    session: Session = ctx.obj["session"]
    if ctx.obj["default_wafer"].name != wafer_name:
        wafer = session.query(Wafer).filter(Wafer.name == wafer_name).first()
    else:
        wafer = ctx.obj["default_wafer"]
    query = (
        session.query(CVMeasurement)
        .filter(CVMeasurement.chip.has(Chip.wafer.__eq__(wafer)))
        .options(joinedload(CVMeasurement.chip))
    )

    # TODO: use EntityChoice with default=ALL instead
    if chips_type is not None:
        query = query.filter(CVMeasurement.chip.has(Chip.type.__eq__(chips_type)))
    else:
        ctx.obj["logger"].info(
            "Chips type (-t or --chips-type) is not specified. Analyzing all chip types."
        )

    query = query.filter(CVMeasurement.chip_state_id.in_((c.id for c in chip_states)))

    if before is not None or after is not None:
        after = after if after is not None else date.min
        before = before if before is not None else date.max
        query = query.filter(CVMeasurement.datetime.between(after, before))

    measurements: list[CVMeasurement] = query.all()

    if not measurements:
        ctx.obj["logger"].warning("No measurements found.")
        return

    chips_types = (
        {chips_type}
        if chips_type is not None
        else {measurement.chip.type for measurement in measurements}
    )

    sheets_data = get_sheets_cv_data(measurements)
    voltages = sorted(Decimal(v) for v in ["-5", "0", "-35"])
    thresholds = get_thresholds(session, "CV")

    file_name = file_name or f"Summary-CV-{wafer_name}"
    file_name = get_indexed_filename(file_name, ("png", "xlsx"))

    if len(chips_types) > 1:
        ctx.obj["logger"].warning(
            f"Multiple chip types are found ({chips_types}). Plotting is not supported and will be skipped."
        )
    else:
        chips_type = next(iter(chips_types))
        fig, axes = plot_data(measurements, voltages, quantile, thresholds.get(chips_type, {}))
        fig.suptitle(wafer_name, fontsize=14)
        for ax_row in axes:
            ax_row[0].set_xlabel("Capacitance [pF]")

        png_file_name = f"{file_name}.png"

        fig.savefig(png_file_name, dpi=300)
        ctx.obj["logger"].info(f"Summary data is plotted to {png_file_name}")

    exel_file_name = f"{file_name}.xlsx"
    info = get_info(wafer=wafer, chip_states=chip_states, measurements=measurements)
    save_cv_summary_to_excel(sheets_data, info, exel_file_name, voltages, thresholds)

    ctx.obj["logger"].info(f"Summary data is saved to {exel_file_name}")


@click.command(name="eqe", help="Make summary for EQE measurements' data.")
@click.pass_context
@click.option(
    "-w",
    "--wafer",
    "wafer_name",
    help="Name of wafer to analyze. Multiple EQE sessions may be found.",
)
@click.option(
    "-o",
    "--output",
    "file_name",
    show_default="Summary-EQE-{wafer/session}",
    help="Output file name without extension.",
)
@click.option(
    "--session",
    "eqe_session",
    help="Date of EQE session or special word 'last'",
    type=click.DateTime(formats=["%Y-%m-%d", "last"]),
    callback=eqe_session_date,
)
@click.option(
    "--no-ref",
    is_flag=True,
    default=False,
    help="Exclude reference measurements from the plots",
)
def summary_eqe(
    ctx: click.Context,
    wafer_name: Optional[str],
    file_name: Optional[str],
    eqe_session: Optional[EqeSession],
    no_ref: bool,
):
    if eqe_session and wafer_name:
        ctx.obj["logger"].error(
            "--wafer and --session options are not allowed to use simultaneously"
        )
        ctx.exit(1)
    elif not eqe_session and not wafer_name:
        ctx.obj["logger"].error("Neither --wafer nor --session are specified")
        ctx.exit(1)

    session: Session = ctx.obj["session"]
    query = (
        session.query(EqeConditions)
        .options(joinedload(EqeConditions.chip))
        .options(joinedload(EqeConditions.measurements))
    )

    if eqe_session:
        query = query.filter(EqeConditions.session == eqe_session)

    if wafer_name:
        eqe_session_ids = session.execute(
            text(
                """
        SELECT DISTINCT eqe_conditions.session_id FROM eqe_conditions
        WHERE eqe_conditions.chip_id IN (
            SELECT chip.id FROM chip WHERE chip.wafer_id = (
                SELECT wafer.id FROM wafer WHERE wafer.name = :wafer_name
            )
        )
        """
            ),
            {"wafer_name": wafer_name},
        ).all()

        if not eqe_session_ids:
            ctx.obj["logger"].warning("No EQE sessions were found for given wafer name")
            ctx.exit()
        query = query.filter(EqeConditions.session_id.in_(chain.from_iterable(eqe_session_ids)))

    conditions = query.all()
    if not conditions:
        ctx.obj["logger"].warning("No measurements found.")
        return

    ctx.obj["logger"].info("EQE data is loaded.")
    sheets_data = get_sheets_eqe_data(conditions)

    file_name = file_name or f"Summary-EQE-{wafer_name or eqe_session.date}"
    file_name = get_indexed_filename(file_name, ("png", "xlsx"))
    png_file_name = f"{file_name}.png"
    exel_file_name = f"{file_name}.xlsx"

    plottable_sheets = [sheet for sheet in sheets_data if sheet.get("prop") is not None]
    fig, axes = plt.subplots(len(plottable_sheets), 1, figsize=(10, 15))
    fig: plt.Figure
    fig.suptitle(
        f"EQE summary for {wafer_name}"
        if wafer_name
        else f"Summary for {eqe_session.date} EQE session"
    )

    for sheet_data, ax in zip(plottable_sheets, axes.ravel()):
        ax: Axes
        last_rows = sheet_data["df"].groupby(by=["Wafer", "Chip"]).last()
        for (wafer_name, chip_name), series in last_rows.iterrows():
            if wafer_name.upper() == "REF" and no_ref:
                continue
            xs = [key for key in series.dropna().keys() if isinstance(key, int)]
            ys = [series[x] for x in xs]
            ax.plot(xs, ys, label=f"{wafer_name} {chip_name}")
            ax.set_xlabel("Wavelength, nm")
            ax.set_ylabel(f"{sheet_data['name']}, {sheet_data['unit']}")
            ax.legend(bbox_to_anchor=(1.05, 1), loc="upper left")
        fig.tight_layout()
    fig.savefig(png_file_name, dpi=300)
    ctx.obj["logger"].info(f"EQE data is plotted to {png_file_name}")

    with pd.ExcelWriter(exel_file_name) as writer:
        for sheet_data in sheets_data:
            sheet_data["df"].to_excel(writer, sheet_name=sheet_data["name"])

    ctx.obj["logger"].info(f"Summary data is saved to {exel_file_name}")


@click.group(
    name="summary",
    help="Group of command to analyze and summarize the data",
    commands=[summary_iv, summary_cv, summary_eqe],
)
@click.pass_context
def summary_group(ctx: click.Context):
    session: Session = ctx.obj["session"]
    active_command = summary_group.commands[ctx.invoked_subcommand]

    try:
        wafer_option = next(
            (o for o in active_command.params if o.name == "wafer_name" and o.prompt)
        )
        last_wafer = session.query(Wafer).order_by(desc(Wafer.record_created_at)).first()
        default_wafer_name = last_wafer.name
        wafer_option.default = default_wafer_name
        ctx.obj["default_wafer"] = last_wafer
    except StopIteration:
        ...


def save_iv_summary_to_excel(
    sheets_data: dict[str, Union[pd.DataFrame, Any]],
    info: pd.Series,
    file_name: str,
    voltages: Iterable[Decimal],
    thresholds: dict[str, dict[Decimal, float]],
):
    summary_df = get_slice_by_voltages(sheets_data["anode"], voltages)
    rules = {
        "lessThan": PatternFill(bgColor="ee9090", fill_type="solid"),
        "greaterThanOrEqual": PatternFill(bgColor="90ee90", fill_type="solid"),
    }

    with pd.ExcelWriter(file_name) as writer:
        summary_df.rename(columns=float).to_excel(writer, sheet_name="Summary")
        apply_conditional_formatting(
            writer.book["Summary"],
            sheets_data["chip_types"],
            rules,
            thresholds,
        )
        I1_anode_df = sheets_data["anode"].dropna(axis=1, how="all").rename(columns=float)
        if not I1_anode_df.empty:
            I1_anode_df.to_excel(writer, sheet_name="I1 anode")
        I3_anode_df = sheets_data["cathode"].dropna(axis=1, how="all").rename(columns=float)
        if not I3_anode_df.empty:
            I3_anode_df.to_excel(writer, sheet_name="I3 anode")
        guard_ring_df = sheets_data["guard_ring"].dropna(axis=1, how="all").rename(columns=float)
        if not guard_ring_df.empty:
            guard_ring_df.to_excel(writer, sheet_name="Guard ring current")
        info.to_excel(writer, sheet_name="Info")


def save_cv_summary_to_excel(
    sheets_data: dict,
    info: pd.Series,
    file_name: str,
    voltages: Iterable[Decimal],
    thresholds: dict[str, dict[Decimal, float]],
):
    summary_df = get_slice_by_voltages(sheets_data["capacitance"], voltages)
    rules = {
        "greaterThanOrEqual": PatternFill(bgColor="ee9090", fill_type="solid"),
        "lessThan": PatternFill(bgColor="90ee90", fill_type="solid"),
    }

    with pd.ExcelWriter(file_name) as writer:
        summary_df.to_excel(writer, sheet_name="Summary")
        apply_conditional_formatting(
            writer.book["Summary"],
            sheets_data["chip_types"],
            rules,
            thresholds,
        )

        sheets_data["capacitance"].rename(columns=float).to_excel(writer, sheet_name="All data")
        info.to_excel(writer, sheet_name="Info")


def get_slice_by_voltages(df: pd.DataFrame, voltages: Iterable[Decimal]) -> pd.DataFrame:
    columns = sorted(voltages)
    slice_df = df[df.columns.intersection(columns)].copy()

    empty_cols = slice_df.isna().all(axis=0)
    if empty_cols.any():
        click.get_current_context().obj["logger"].warning(
            f"""The following voltages are not present in the data: {
            [float(col) for col, val in empty_cols.items() if val]}."""
        )
    slice_df.dropna(axis=1, how="all", inplace=True)
    return slice_df


def apply_conditional_formatting(
    sheet: Worksheet,
    chip_types: list[str],
    rules: dict[str, Fill],
    thresholds: dict[str, dict[Decimal, float]],
):
    chip_row_index = [(i + 1, cell.value) for i, cell in enumerate(sheet["A"]) if cell.value]

    for chip_type in chip_types:

        def is_current_type(chip_name: str) -> bool:
            return chip_name.startswith(chip_type)

        chip_type_thresholds = thresholds.get(chip_type)
        if chip_type_thresholds is None:
            continue
        for voltage, threshold in chip_type_thresholds.items():
            try:
                column_cell = next(
                    (
                        cell
                        for cell in sheet["1"]
                        if cell.value is not None and Decimal(str(cell.value)) == voltage
                    )
                )
                first_row_index = next(i for i, v in chip_row_index if is_current_type(v))
                last_row_index = next(i for i, v in reversed(chip_row_index) if is_current_type(v))
            except ValueError:
                continue
            except StopIteration:
                continue

            column_letter = column_cell.column_letter
            cell_range = f"{column_letter}{first_row_index}:{column_letter}{last_row_index}"
            for rule_name, rule in rules.items():
                cells_rule = CellIsRule(operator=rule_name, formula=[threshold], fill=rule)
                sheet.conditional_formatting.add(cell_range, cells_rule)


def get_sheets_iv_data(
    conditions: list[IvConditions],
) -> dict[str, Union[pd.DataFrame, Any]]:
    chip_names = sorted({condition.chip.name for condition in conditions})
    chip_types = {condition.chip.type for condition in conditions}
    anode_df = pd.DataFrame(dtype="float64", index=chip_names)
    cathode_df = pd.DataFrame(dtype="float64", index=chip_names)
    guard_ring_df = pd.DataFrame(dtype="float64", index=chip_names)
    has_uncorrected_current = False
    with click.progressbar(conditions, label="Processing measurements...") as progress:
        for condition in progress:
            for measurement in condition.measurements:
                measurement: IVMeasurement
                cell_location = (condition.chip.name, measurement.voltage_input)
                if measurement.anode_current_corrected is None:
                    has_uncorrected_current = True
                anode_df.loc[cell_location] = measurement.get_anode_current_value()
                cathode_df.loc[cell_location] = measurement.cathode_current
                guard_ring_df.loc[cell_location] = measurement.guard_current
    if has_uncorrected_current:
        click.get_current_context().obj["logger"].warning(
            "Some current measurements are not corrected by temperature."
        )
    return {
        "anode": anode_df,
        "cathode": cathode_df,
        "guard_ring": guard_ring_df,
        "chip_names": chip_names,
        "chip_types": chip_types,
    }


def get_sheets_cv_data(
    measurements: list[CVMeasurement],
) -> dict[str, Union[pd.DataFrame, Any]]:
    chip_names = sorted({measurement.chip.name for measurement in measurements})
    chip_types = {measurement.chip.type for measurement in measurements}
    voltages = sorted({measurement.voltage_input for measurement in measurements})
    capacitance_df = pd.DataFrame(dtype="float64", index=chip_names, columns=voltages)
    with click.progressbar(measurements, label="Processing measurements...") as progress:
        for measurement in progress:
            measurement: CVMeasurement
            cell_location = (measurement.chip.name, measurement.voltage_input)
            capacitance_df.loc[cell_location] = measurement.capacitance
    return {
        "capacitance": capacitance_df,
        "chip_names": chip_names,
        "chip_types": chip_types,
        "voltages": voltages,
    }


def get_sheets_eqe_data(conditions: list[EqeConditions]) -> list[dict]:
    all_sheets = []

    series_list = []
    for condition in conditions:
        series_list.append(
            pd.Series(
                [
                    condition.datetime,
                    condition.chip.wafer.name,
                    condition.chip.name,
                    condition.bias,
                    condition.averaging,
                    condition.dark_current,
                    condition.temperature,
                ],
                [
                    "Datetime",
                    "Wafer",
                    "Chip",
                    "Bias",
                    "Averaging",
                    "Dark current",
                    "Temperature",
                ],
            )
        )
    df_info = pd.concat(series_list, axis=1, ignore_index=True).T
    df_info.set_index("Datetime", inplace=True)
    df_info = df_info.sort_index()
    all_sheets.append({"df": df_info, "name": "Info"})

    for prop, name, unit in (
        ("eqe", "EQE", "%"),
        ("light_current", "Light current", "A"),
        ("dark_current", "Dark current", "A"),
        ("responsivity", "Responsivity", "A/W"),
        ("std", "Standard deviation", "A"),
    ):
        series_list = []
        for condition in conditions:
            measurements = sorted(condition.measurements, key=lambda m: m.wavelength)
            series = pd.Series(
                [
                    condition.datetime,
                    condition.chip.wafer.name,
                    condition.chip.name,
                    *(getattr(m, prop) for m in measurements),
                ],
                index=[
                    "Datetime",
                    "Wafer",
                    "Chip",
                    *(m.wavelength for m in measurements),
                ],
            )
            series_list.append(series)
        df = pd.concat(series_list, axis=1, ignore_index=True).T
        if df.dropna(axis=1, how="all").empty:
            continue
        df.set_index("Datetime", inplace=True)
        df = df.sort_index()
        all_sheets.append({"name": name, "df": df, "prop": prop, "unit": unit})

    return all_sheets


def get_info(
    wafer: Wafer,
    chip_states: Iterable[ChipState],
    measurements: list[Union[IvConditions, CVMeasurement]],
) -> pd.Series:
    format_date = strftime("%A, %d %b %Y", localtime())
    chip_states_str = "; ".join([state.name for state in chip_states])

    first_measurement = min(measurements, key=lambda m: m.datetime)
    last_measurement = max(measurements, key=lambda m: m.datetime)

    return pd.Series(
        {
            "Wafer": wafer.name,
            "Summary generation date": format_date,
            "Chip state": chip_states_str,
            "First measurement date": first_measurement.datetime,
            "Last measurement date": last_measurement.datetime,
        }
    )


def get_indexed_filename(filename, extensions):
    index = max((sum((1 for f in Path(".").glob(f"{filename}*.{ext}"))) for ext in extensions)) + 1
    if index > 1:
        filename = f"{filename}-{index}"
    return filename


def plot_hist(ax: Axes, data: np.ndarray):
    ax.set_ylabel("Number of chips")
    ax.hist(data * 1e12, bins=15)


def plot_heat_map(ax: Axes, values: np.ndarray, xs: np.ndarray, ys: np.ndarray, vmin, vmax):
    width = xs.max() - xs.min() + 1
    height = ys.max() - ys.min() + 1
    grid = np.full((height, width), np.nan)
    grid[ys - ys.min(), xs - xs.min()] = values

    X = np.linspace(min(xs) - 0.5, max(xs) + 0.5, width + 1)
    Y = np.linspace(min(ys) - 0.5, max(ys) + 0.5, height + 1)
    mesh = ax.pcolormesh(X, Y, grid, cmap="hot", shading="flat", vmin=vmin, vmax=vmax)
    ax.xaxis.set_major_locator(MaxNLocator(integer=True, min_n_ticks=0))
    ax.yaxis.set_major_locator(MaxNLocator(integer=True, min_n_ticks=0))
    ax.set_ylabel("Y coordinate")
    ax.set_xlabel("X coordinate")
    ax.figure.colorbar(mesh, ax=ax)


def get_iv_plot_data(measurements: Collection[IVMeasurement]):
    data = np.array([value.get_anode_current_value() for value in measurements])
    xs = np.array([measurement.conditions.chip.x_coordinate for measurement in measurements])
    ys = np.array([measurement.conditions.chip.y_coordinate for measurement in measurements])
    return data, xs, ys


def get_cv_plot_data(measurements: Collection[CVMeasurement]):
    data = np.array([value.capacitance for value in measurements])
    xs = np.array([measurement.chip.x_coordinate for measurement in measurements])
    ys = np.array([measurement.chip.y_coordinate for measurement in measurements])
    return data, xs, ys


def plot_data(
    values: Collection[Union[IVMeasurement, CVMeasurement]],
    voltages: Collection[Decimal],
    quantile: tuple[float, float],
    thresholds: dict[Decimal, float],
) -> (Figure, ndarray[Any, Axes]):
    failure_map = np.isclose(quantile, [0, 0], atol=1e-5).all()
    cols_num = 1 if failure_map else 2
    fig, axes = plt.subplots(
        nrows=len(voltages),
        ncols=cols_num,
        figsize=(5 * cols_num, 5 * len(voltages)),
        gridspec_kw={
            "left": 0.1,
            "right": 0.95,
            "bottom": 0.05,
            "top": 0.95,
            "wspace": 0.3,
            "hspace": 0.35,
        },
    )
    axes = axes.reshape(-1, cols_num)

    with click.progressbar(tuple(enumerate(sorted(voltages))), label="Plotting...") as progress:
        for i, voltage in progress:
            target_values = [value for value in values if value.voltage_input == voltage]
            if len(target_values) == 0:
                continue
            if isinstance(target_values[0], IVMeasurement):
                data, xs, ys = get_iv_plot_data(target_values)
            elif isinstance(target_values[0], CVMeasurement):
                data, xs, ys = get_cv_plot_data(target_values)
            else:
                raise RuntimeError("Unknown object type was provided")

            axes[i][0].set_title(f"{voltage}V")

            if failure_map:
                if voltage not in thresholds:
                    click.get_current_context().obj["logger"].warning(
                        f"Thresholds for {voltage}V are not found. Skipping."
                    )
                    continue
                clipped_data = data > thresholds[voltage]
                lower_bound, upper_bound = -1, 1.2
                plot_heat_map(axes[i][0], clipped_data, xs, ys, lower_bound, upper_bound)
            else:
                axes[i][1].set_title(f"{voltage}V")
                lower_bound, upper_bound = np.quantile(data, quantile)
                clipped_data = np.clip(data, lower_bound, upper_bound)
                plot_hist(axes[i][0], clipped_data)
                plot_heat_map(axes[i][1], clipped_data, xs, ys, lower_bound, upper_bound)
    return fig, axes
