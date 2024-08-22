from datetime import (
    date,
    datetime,
)
from decimal import Decimal
from typing import (
    Iterable,
    TypedDict,
)

import click
import pandas as pd
from openpyxl.styles import PatternFill
from pandas import DataFrame
from sqlalchemy.orm import (
    Query,
    joinedload,
)

from orm import (
    CVMeasurement,
    Chip,
    ChipState,
    Wafer,
)
from utils import (
    EntityOption,
    get_indexed_filename,
    get_thresholds,
    validate_chip_types,
    wafer_loader,
)
from .common import (
    apply_conditional_formatting,
    date_formats,
    date_formats_help,
    get_info,
    get_slice_by_voltages,
    plot_data,
)
from ..context import (
    AnalyzerContext,
    pass_analyzer_context,
)


class SheetsCVData(TypedDict):
    capacitance: DataFrame
    chip_names: list[str]
    voltages: list[Decimal]


@click.command(name="cv", help="Make summary (png and xlsx) for CV measurements' data.")
@pass_analyzer_context
@click.option(
    "-t", "--chips-type", help="Type of the chips to analyze.",
    callback=validate_chip_types,
)
@click.option(
    "-w", "--wafer", prompt="Wafer name", help="Wafer name.",
    required=True,
    callback=wafer_loader,
)
@click.option(
    "-s",
    "--chip-state",
    "chip_states",
    help="State of the chips to summarize.",
    default=["ALL"],
    show_default=True,
    multiple=True,
    cls=EntityOption,
    entity_type=ChipState,
)
@click.option(
    "-q",
    "--quantile",
    default=(0.01, 0.99),
    show_default=True,
    type=click.Tuple([float, float]),
    help="Min and max plotted values cutoff.",
)
@click.option(
    "--before",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements before (exclusive) provided date and time. {date_formats_help}",
)
@click.option(
    "--after",
    type=click.DateTime(formats=date_formats),
    help=f"Include measurements after (inclusive) provided date and time. {date_formats_help}",
)
def summary_cv(
    ctx: AnalyzerContext,
    chips_type: str | None,
    wafer: Wafer,
    chip_states: list[ChipState],
    quantile: tuple[float, float],
    before: datetime | date | None,
    after: datetime | date | None,
):
    query: Query = (
        ctx.session.query(CVMeasurement)
        .filter(CVMeasurement.chip.has(Chip.wafer == wafer))
        .options(joinedload(CVMeasurement.chip))
    )
    
    if chips_type is not None:
        query = query.filter(CVMeasurement.chip.has(Chip.type == chips_type))
    else:
        ctx.logger.info(
            "Chips type (-t or --chips-type) is not specified. Analyzing all chip types."
        )
    
    query = query.filter(CVMeasurement.chip_state_id.in_((c.id for c in chip_states)))
    
    if before is not None or after is not None:
        after = after if after is not None else date.min
        before = before if before is not None else date.max
        query = query.filter(CVMeasurement.datetime.between(after, before))
    
    measurements: list[CVMeasurement] = query.all()
    
    if not measurements:
        ctx.logger.warning("No measurements found.")
        return
    
    chips_types = (
        {chips_type}
        if chips_type is not None
        else {measurement.chip.type for measurement in measurements}
    )
    
    sheets_data = get_sheets_cv_data(measurements)
    voltages = sorted(Decimal(v) for v in ["-5", "0", "-35"])
    thresholds = get_thresholds(ctx.session, "CV")
    
    file_name = get_indexed_filename(f"Summary-CV-{wafer.name}", ("png", "xlsx"))
    
    if len(chips_types) > 1:
        ctx.logger.warning(
            f"Multiple chip types are found ({chips_types}). "
            "Plotting is not supported and will be skipped.")
    else:
        chips_type = next(iter(chips_types))
        fig, axes = plot_data(measurements, voltages, quantile, thresholds.get(chips_type, {}))
        fig.suptitle(wafer.name, fontsize=14)
        
        png_file_name = f"{file_name}.png"
        
        fig.savefig(png_file_name, dpi=300)
        ctx.logger.info(f"Summary data is plotted to {png_file_name}")
    
    exel_file_name = f"{file_name}.xlsx"
    info = get_info(wafer=wafer, chip_states=chip_states, measurements=measurements)
    save_cv_summary_to_excel(sheets_data, info, exel_file_name, voltages, thresholds)
    
    ctx.logger.info(f"Summary data is saved to {exel_file_name}")


def save_cv_summary_to_excel(
    sheets_data: SheetsCVData,
    info: pd.Series,
    file_name: str,
    voltages: Iterable[Decimal],
    thresholds: dict[str, dict[Decimal, float]],
):
    summary_df = get_slice_by_voltages(sheets_data["capacitance"], voltages)
    rules = {
        "greaterThanOrEqual": PatternFill(bgColor="ee9090", fill_type="solid"),
        "lessThan": PatternFill(bgColor="90ee90", fill_type="solid"),
    }
    
    with pd.ExcelWriter(file_name) as writer:
        summary_df.to_excel(writer, sheet_name="Summary")
        apply_conditional_formatting(
            writer.book["Summary"],
            rules,
            thresholds,
        )
        
        sheets_data["capacitance"].rename(columns=float).to_excel(writer, sheet_name="All data")
        info.to_excel(writer, sheet_name="Info")


def get_sheets_cv_data(
    measurements: list[CVMeasurement],
) -> SheetsCVData:
    chip_names: list[str] = sorted({measurement.chip.name for measurement in measurements})
    voltages: list[Decimal] = sorted({measurement.voltage_input for measurement in measurements})
    capacitance_df = pd.DataFrame(index=pd.Index(chip_names), columns=pd.Index(voltages))
    with click.progressbar(measurements, label="Processing measurements...") as progress:
        for measurement in progress:
            measurement: CVMeasurement
            cell_location = (measurement.chip.name, measurement.voltage_input)
            capacitance_df.loc[cell_location] = measurement.capacitance
    return {
        "capacitance": capacitance_df,
        "chip_names": chip_names,
        "voltages": voltages,
    }
